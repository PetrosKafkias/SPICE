import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"
APP_DATA = ROOT / "src" / "app" / "data"
APPROVED_STATUSES = {"Formatted for printing", "Content ready"}
PARTNERS = {"CERTH", "KEPA", "DReVen", "CdT", "IRID", "UAB", "ENSAN", "IHNet", "EURADA", "miB"}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\u200d", "").replace("_x0003_", "").strip()
    return re.sub(r"\r\n?", "\n", text)


def slugify(value):
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")


def rows_with_values(sheet):
    return [
        list(row)
        for row in sheet.iter_rows(values_only=True)
        if any(value not in (None, "") for value in row)
    ]


def phase_from_scope(scope):
    match = re.search(r"\b([1-5])\s*\|", scope)
    return int(match.group(1)) if match else 1


def mode_from_source(value):
    source = value.lower()
    if "online" in source and "offline" in source:
        return "Hybrid"
    if "online" in source:
        return "Online"
    return "Offline"


def split_tags(value):
    return [part.strip() for part in re.split(r",|\n", value) if part.strip()]


def extract_tools():
    workbook = load_workbook(DOWNLOADS / "Participatory tools and methods.xlsx", data_only=True)
    sheet = workbook["🛠️ Tools and methods"]
    rows = rows_with_values(sheet)
    headers = [clean(value) for value in rows[0]]
    index = {header: position for position, header in enumerate(headers) if header}

    def cell(row, header):
        position = index.get(header)
        return clean(row[position]) if position is not None and position < len(row) else ""

    tools = []
    used_ids = set()
    for row in rows[1:]:
        name = cell(row, "NAME")
        status = cell(row, "Status")
        if not name or status not in APPROVED_STATUSES:
            continue
        base_id = slugify(name)
        tool_id = base_id
        suffix = 2
        while tool_id in used_ids:
            tool_id = f"{base_id}-{suffix}"
            suffix += 1
        used_ids.add(tool_id)

        scope = cell(row, "Scope")
        phase = phase_from_scope(scope)
        description = cell(row, "Description")
        short_description = cell(row, "How to use (short, for preview)") or description
        facilitator = cell(row, "Facilitator/participant")
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", facilitator):
            facilitator = "Not specified"

        tools.append(
            {
                "id": tool_id,
                "name": name,
                "status": status,
                "typology": cell(row, "Typology"),
                "shortDesc": short_description,
                "purpose": description,
                "phase": phase,
                "phaseName": scope.split("|", 1)[1].strip().title() if "|" in scope else "Framing & Readiness",
                "objectiveTags": split_tags(cell(row, "Ambition")),
                "mode": mode_from_source(cell(row, "Recommended Mode")),
                "duration": cell(row, "Reccomended time") or "Not specified",
                "groupSize": "Not specified",
                "budget": "Flexible",
                "facilitatorRatio": facilitator or "Not specified",
                "suppliesRequired": cell(row, "Supplies required"),
                "expectedOutputs": [],
                "accessibilityNotes": "",
                "usageTip": cell(row, "Tips for facilitation"),
                "proTip": "",
                "implementationTime": cell(row, "Reccomended time") or "Not specified",
                "developmentTime": "Not specified",
                "howTo": cell(row, "How to use (complete)"),
                "budgetAdaptation": cell(row, "Adaptation according to budget"),
                "examples": cell(row, "Examples"),
                "requirements": cell(row, "Requirements"),
                "reference": cell(row, "Reference "),
                "printableUrl": cell(row, "Link to the printable version "),
                "onlineResources": cell(row, "ONLINE resources (for platform)"),
            }
        )
    return tools


def extract_glossary():
    workbook = load_workbook(DOWNLOADS / "Glossary.xlsx", data_only=True)
    entries = []
    seen = set()
    for sheet in workbook.worksheets:
        rows = rows_with_values(sheet)
        if not rows:
            continue
        headers = [clean(value) for value in rows[0]]
        if "Words" not in headers:
            continue
        term_index = headers.index("Words")
        relevance_index = next((i for i, header in enumerate(headers) if "Relevance" in header or header == "Column 13"), None)
        category_index = 0 if headers[0] not in {"", "Words", "Column 14"} else None
        definition_indexes = [i for i, header in enumerate(headers) if header in PARTNERS]

        for row in rows[1:]:
            term = clean(row[term_index]) if term_index < len(row) else ""
            if not term:
                continue
            key = term.casefold()
            if key in seen:
                continue
            seen.add(key)
            definitions = []
            for position in definition_indexes:
                value = clean(row[position]) if position < len(row) else ""
                if value and value != "-":
                    definitions.append({"owner": headers[position], "text": value})
            category = (
                clean(row[category_index])
                if category_index is not None and category_index < len(row)
                else ("Digital Tools" if sheet.title == "Technology and digital tools" else "Co-Creation Process")
            )
            if not category:
                category = "Co-Creation Process"
            entries.append(
                {
                    "id": slugify(term),
                    "term": term,
                    "definition": definitions[0]["text"] if definitions else None,
                    "definitionOwner": definitions[0]["owner"] if definitions else None,
                    "category": category.title() if category.isupper() else category,
                    "language": "en",
                    "relevance": clean(row[relevance_index]) if relevance_index is not None and relevance_index < len(row) else "",
                    "sourceSheet": sheet.title,
                }
            )
    return entries


APP_DATA.mkdir(parents=True, exist_ok=True)
(APP_DATA / "analogueToolsSource.json").write_text(
    json.dumps(extract_tools(), ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
)
(APP_DATA / "glossarySource.json").write_text(
    json.dumps(extract_glossary(), ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
)
